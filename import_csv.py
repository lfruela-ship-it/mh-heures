#!/usr/bin/env python3
"""
import_csv.py  v2.0 — MH Maler & Gipser AG
============================================
Importe un CSV (format MH_IMPORT_V2) généré par la PWA
dans le fichier MH_Gestion_Heures_2026_PRO.xlsx

USAGE:
  python import_csv.py <fichier.csv> [fichier_excel.xlsx]

NOUVEAUTÉS v2.0:
  - Support multi-chantiers (colonne CHANTIERS JSON)
  - Détection automatique de la version CSV
  - Backup automatique avant import
  - Rapport détaillé des données importées
"""

import sys, os, csv, datetime, shutil, json
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font
except ImportError:
    print("❌  pip install openpyxl"); sys.exit(1)

VERSION = "v2.0"

SHEET_MAP = {
    1:'Jan',2:'Fév',3:'Mar',4:'Avr',5:'Mai',6:'Jun',
    7:'Jul',8:'Aoû',9:'Sep',10:'Oct',11:'Nov',12:'Déc'
}
MOIS_FR = {
    'Janvier':1,'Février':2,'Mars':3,'Avril':4,'Mai':5,'Juin':6,
    'Juillet':7,'Août':8,'Septembre':9,'Octobre':10,'Novembre':11,'Décembre':12
}

FILL_IMP  = PatternFill("solid", fgColor="EAF4FB")
FILL_MULT = PatternFill("solid", fgColor="E8F5E9")  # multi-chantiers
FILL_CODE = PatternFill("solid", fgColor="FFF9C4")

def t2excel(t_str):
    if not t_str or not t_str.strip(): return None
    try:
        h,m = map(int, t_str.strip().split(':'))
        return (h*60+m)/(24*60)
    except: return None

def find_row(ws, date_iso):
    try:
        target = datetime.date.fromisoformat(date_iso)
    except: return None
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        val = row[0].value
        if val is None: continue
        if isinstance(val, (datetime.datetime, datetime.date)):
            d = val.date() if isinstance(val, datetime.datetime) else val
            if d == target: return row[0].row
        elif isinstance(val, (int,float)):
            try:
                d = datetime.date(1899,12,30) + datetime.timedelta(days=int(val))
                if d == target: return row[0].row
            except: pass
    return None

def import_csv(csv_path, excel_path):
    print(f"\n{'='*56}")
    print(f"  MH Maler & Gipser AG — Import CSV {VERSION}")
    print(f"{'='*56}")
    print(f"  CSV   : {csv_path}")
    print(f"  Excel : {excel_path}\n")

    with open(csv_path,'r',encoding='utf-8-sig') as f:
        raw = f.read()

    # Parse metadata
    employe=periode=None; annee=mois_num=0
    for line in raw.split('\n'):
        line=line.strip()
        if line.startswith('## EMPLOYE:'): employe=line[11:].strip()
        elif line.startswith('## PERIODE:'):
            p=line[11:].strip()
            if len(p)==7:
                try: annee,mois_num=int(p[:4]),int(p[5:])
                except: pass
        elif line.startswith('## ANNEE:') and not annee:
            try: annee=int(line[9:].strip())
            except: pass
        elif line.startswith('## MOIS:') and not mois_num:
            mois_num=MOIS_FR.get(line[8:].strip(),0)

    # Fallback from data
    lines=[l for l in raw.split('\n') if not l.startswith('#') and l.strip()]
    if not lines: print("❌ Aucune donnée."); sys.exit(1)
    reader=csv.DictReader(lines); rows=list(reader)
    if not rows: print("❌ CSV vide."); sys.exit(1)
    if not (annee and mois_num):
        try:
            d=datetime.date.fromisoformat(rows[0].get('DATE_ISO','').strip())
            annee=annee or d.year; mois_num=mois_num or d.month
        except: pass

    if not (annee and mois_num): print("❌ Mois/année indétectables."); sys.exit(1)

    sheet_name=SHEET_MAP.get(mois_num)
    if not sheet_name: print(f"❌ Mois invalide: {mois_num}"); sys.exit(1)

    print(f"  Employé  : {employe or '—'}")
    print(f"  Période  : {mois_num:02d}/{annee}")
    print(f"  Feuille  : {sheet_name}")
    print(f"  Lignes   : {len(rows)}\n")

    if not Path(excel_path).exists():
        print(f"❌ Excel introuvable: {excel_path}"); sys.exit(1)

    wb=load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        print(f"❌ Onglet '{sheet_name}' non trouvé."); sys.exit(1)
    ws=wb[sheet_name]

    ok=skip=err=0
    for row_data in rows:
        date_iso=row_data.get('DATE_ISO','').strip()
        if not date_iso: skip+=1; continue
        excel_row=find_row(ws,date_iso)
        if not excel_row: err+=1; print(f"  ⚠ {date_iso} non trouvé"); continue

        is_sat=datetime.date.fromisoformat(date_iso).weekday()==5
        fill=FILL_IMP

        # Heures
        for key,col in [('DEB1',3),('FIN1',4),('DEB2',5),('FIN2',6),('DEB3',7),('FIN3',8)]:
            v=t2excel(row_data.get(key,''))
            c=ws.cell(excel_row,col)
            c.value=v; c.fill=fill
            if v: c.number_format='H:MM'

        # Code
        code=row_data.get('CODE','').strip()
        cc=ws.cell(excel_row,10)
        cc.value=code or None
        if code: cc.fill=FILL_CODE; cc.font=Font(bold=True,color='E8532B')
        else: cc.fill=fill

        # Commentaire (K=11)
        ws.cell(excel_row,11).value=row_data.get('COMMENTAIRE','').strip() or None
        ws.cell(excel_row,11).fill=fill

        # Multi-chantiers ou chantier simple
        chantiers_json=row_data.get('CHANTIERS','').strip()
        if chantiers_json and chantiers_json not in ('', '[]'):
            try:
                chs=json.loads(chantiers_json)
                if chs:
                    # Tous les chantiers → colonne L, séparés par virgule
                    noms=[c.get('name','').strip() for c in chs if c.get('name','').strip()]
                    ws.cell(excel_row,12).value=', '.join(noms)
                    ws.cell(excel_row,12).fill=FILL_MULT
                    from openpyxl.styles import Alignment
                    ws.cell(excel_row,12).alignment=Alignment(horizontal='left',vertical='center',wrap_text=True)
                    # Km total → colonne N
                    km_tot=sum(c.get('km',0) for c in chs)
                    if km_tot: ws.cell(excel_row,14).value=km_tot
            except: pass
        else:
            chantier=row_data.get('CHANTIER','').strip()
            ws.cell(excel_row,12).value=chantier or None
            ws.cell(excel_row,12).fill=fill
            km_str=row_data.get('KM','').strip()
            if km_str and km_str!='0':
                try: ws.cell(excel_row,14).value=int(km_str)
                except: pass

        ok+=1
        sig=f"  ✓ {date_iso} → ligne {excel_row}"
        if code: sig+=f" [{code}]"
        print(sig)

    # Backup + save
    backup=str(excel_path).replace('.xlsx','_BACKUP.xlsx').replace('.xlsm','_BACKUP.xlsm')
    shutil.copy2(excel_path,backup)
    wb.save(excel_path)

    print(f"\n{'─'*56}")
    print(f"  ✅ Import terminé  —  {ok} jours importés")
    if skip: print(f"  Ignorés  : {skip}")
    if err:  print(f"  Erreurs  : {err}")
    print(f"  Backup   : {Path(backup).name}")
    print(f"  → Ctrl+Alt+F9 dans Excel pour recalculer\n")

if __name__=='__main__':
    if len(sys.argv)<2: print(__doc__); sys.exit(0)
    csv_path=sys.argv[1]
    if len(sys.argv)>=3:
        excel_path=sys.argv[2]
    else:
        candidates=[c for c in Path('.').glob('MH_Gestion_Heures*.xls*') if 'BACKUP' not in str(c)]
        if candidates: excel_path=str(candidates[0]); print(f"ℹ Excel: {excel_path}")
        else: print("❌ Excel non trouvé. Spécifiez-le en argument."); sys.exit(1)
    import_csv(csv_path,excel_path)
